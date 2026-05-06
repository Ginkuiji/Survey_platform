from io import BytesIO
from typing import Any

from django.utils import timezone


def format_value(value: Any):
    if value is None or value == "":
        return "—"
    return value


def format_number(value: Any):
    if value is None or value == "":
        return "—"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    return round(number, 4)


def format_p_value(value: Any):
    if value is None or value == "":
        return "—"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    if 0 < number < 0.0001:
        return "<0.0001"
    return round(number, 4)


def format_datetime(value: Any) -> str:
    if not value:
        return "—"
    try:
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime("%d.%m.%Y %H:%M")
    except AttributeError:
        return str(value)


def get_variable_label(result: dict, code: str) -> str:
    if not code:
        return "—"
    if code == "intercept":
        return "Свободный член"
    variable = (result.get("variables_by_code") or {}).get(code) or {}
    return variable.get("label") or code


def append_key_value(ws, key, value):
    ws.append([key, format_value(value)])


def append_section_title(ws, title):
    row_number = ws.max_row + 1
    ws.append([title])
    ws.cell(row=row_number, column=1).style = "Title"


def autosize_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        ws.column_dimensions[column_letter].width = max(max_length + 2, 12)


def _style_header(ws, row_number):
    from openpyxl.styles import Font, PatternFill

    for cell in ws[row_number]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EDEDED")


def _append_table(ws, headers, rows):
    start = ws.max_row + 1
    ws.append(headers)
    _style_header(ws, start)
    for row in rows:
        ws.append([format_value(value) for value in row])


def _report_result(analysis_report) -> dict:
    result = analysis_report.result or {}
    return result if isinstance(result, dict) else {}


def _ordered_questions(survey, questions):
    order_by_id = {
        question.id: index
        for index, question in enumerate(
            survey.questions.select_related("page").order_by("page__order", "order", "id")
        )
    }
    return sorted(questions, key=lambda item: order_by_id.get(item.get("id"), 10**9))


def _add_questions_sheet(ws, survey, analytic_data):
    questions = _ordered_questions(survey, analytic_data.get("questions") or [])
    for question in questions:
        result = question.get("result") or {}
        base = question.get("base") or {}
        qtype = question.get("qtype")
        append_section_title(ws, f"Вопрос {question.get('id')}")
        append_key_value(ws, "ID вопроса", question.get("id"))
        append_key_value(ws, "Текст", question.get("text"))
        append_key_value(ws, "Тип", qtype)
        append_key_value(ws, "Завершили опрос", base.get("total_completed"))
        append_key_value(ws, "Видели вопрос", base.get("shown_count"))
        append_key_value(ws, "Ответили", base.get("answered_count"))
        append_key_value(ws, "Пропустили", base.get("skipped_count"))

        if qtype in ("single", "multi", "dropdown", "yesno"):
            _append_table(
                ws,
                ["Вариант", "Количество", "% от ответивших", "% от завершивших"],
                [
                    [
                        option.get("text"),
                        option.get("count"),
                        format_number(option.get("percent_answered")),
                        format_number(option.get("percent_total")),
                    ]
                    for option in result.get("options") or []
                ],
            )
        elif qtype in ("scale", "number"):
            append_key_value(ws, "Среднее", format_number(result.get("average")))
            append_key_value(ws, "Медиана", format_number(result.get("median")))
            append_key_value(ws, "Минимум", format_number(result.get("min")))
            append_key_value(ws, "Максимум", format_number(result.get("max")))
            _append_table(
                ws,
                ["Значение", "Количество"],
                [[item.get("value"), item.get("count")] for item in result.get("distribution") or []],
            )
        elif qtype in ("text", "date"):
            append_key_value(ws, "Количество текстовых ответов", result.get("total_text_answers"))
            _append_table(
                ws,
                ["№", "Ответ"],
                [[index + 1, text] for index, text in enumerate((result.get("text_answers") or [])[:20])],
            )
        elif qtype == "matrix_single":
            rows = []
            for matrix_row in result.get("rows") or []:
                for column in matrix_row.get("columns") or []:
                    rows.append([
                        matrix_row.get("text"),
                        column.get("text"),
                        column.get("count"),
                        format_number(column.get("percent_answered")),
                        format_number(column.get("percent_total")),
                    ])
            _append_table(ws, ["Строка", "Колонка", "Количество", "% от ответивших", "% от завершивших"], rows)
        elif qtype == "matrix_multi":
            append_section_title(ws, "Matrix")
            matrix_rows = result.get("rows") or []
            matrix_columns = result.get("columns") or []
            cells_by_key = {
                (cell.get("row_id"), cell.get("column_id")): cell
                for cell in result.get("cells") or []
            }
            ws.append(["Row"] + [column.get("text") for column in matrix_columns])
            _style_header(ws, ws.max_row)
            for matrix_row in matrix_rows:
                ws.append([
                    matrix_row.get("text"),
                    *[
                        (
                            f"{cell.get('count', 0)} | "
                            f"{format_number(cell.get('percent_answered'))}% answered | "
                            f"{format_number(cell.get('percent_total'))}% total"
                        )
                        for column in matrix_columns
                        for cell in [cells_by_key.get((matrix_row.get("id"), column.get("id")), {})]
                    ],
                ])
            _append_table(
                ws,
                ["Row Summary", "Selected total", "Respondents", "% respondents", "Avg selected"],
                [
                    [
                        item.get("row_text"),
                        item.get("selected_total"),
                        item.get("respondent_count"),
                        format_number(item.get("respondent_share")),
                        format_number(item.get("avg_selected_per_respondent")),
                    ]
                    for item in result.get("row_summary") or []
                ],
            )
            _append_table(
                ws,
                ["Column Summary", "Selected total", "Respondents", "% respondents"],
                [
                    [
                        item.get("column_text"),
                        item.get("selected_total"),
                        item.get("respondent_count"),
                        format_number(item.get("respondent_share")),
                    ]
                    for item in result.get("column_summary") or []
                ],
            )
        elif qtype == "ranking":
            options = result.get("options") or []
            _append_table(
                ws,
                ["Вариант", "Средний ранг", "Первое место"],
                [[option.get("text"), format_number(option.get("average_rank")), option.get("first_place_count")] for option in options],
            )
            rank_rows = []
            for option in options:
                for item in option.get("rank_distribution") or []:
                    rank_rows.append([
                        option.get("text"),
                        item.get("rank"),
                        item.get("count"),
                        format_number(item.get("percent_answered")),
                        format_number(item.get("percent_total")),
                    ])
            _append_table(ws, ["Вариант", "Ранг", "Количество", "% от ответивших", "% от завершивших"], rank_rows)
        ws.append([])


def _append_matrix(ws, title, variables, matrix, formatter=format_number):
    append_section_title(ws, title)
    labels = [variable.get("label") or variable.get("code") for variable in variables]
    if not labels and matrix:
        labels = [f"C{index + 1}" for index in range(len(matrix[0]))]
    ws.append([""] + labels)
    _style_header(ws, ws.max_row)
    for index, row in enumerate(matrix or []):
        label = labels[index] if index < len(labels) else f"R{index + 1}"
        ws.append([label] + [formatter(value) for value in row])


def _append_crosstab(ws, crosstab):
    rows = crosstab.get("rows") or []
    if not rows:
        ws.append(["Нет данных для таблицы сопряжённости"])
        return
    column_values = [column.get("value") for column in rows[0].get("columns") or []]
    ws.append([crosstab.get("row_variable") or "row"] + column_values + ["Итого"])
    _style_header(ws, ws.max_row)
    for row in rows:
        ws.append([
            row.get("value"),
            *[
                f"{column.get('count')} | row {format_number(column.get('percent_row'))}% | total {format_number(column.get('percent_total'))}%"
                for column in row.get("columns") or []
            ],
            row.get("total"),
        ])


def _add_report_sheet(ws, analysis_report):
    report_result = _report_result(analysis_report)
    append_key_value(ws, "Название отчёта", analysis_report.title)
    append_key_value(ws, "Дата создания", format_datetime(analysis_report.created_at))
    append_key_value(ws, "Количество секций", len(report_result.get("sections") or []))
    ws.append([])

    for section in report_result.get("sections") or []:
        title = section.get("title") or section.get("type")
        section_type = section.get("type")
        result = section.get("result") or {}
        append_section_title(ws, title)
        append_key_value(ws, "Тип", section_type)

        if section.get("error"):
            append_key_value(ws, "Ошибка", section.get("error"))
            ws.append([])
            continue

        if section_type == "correlation":
            append_key_value(ws, "Метод", result.get("method"))
            append_key_value(ws, "Dataset size", result.get("dataset_size"))
            variables = result.get("variables") or []
            _append_matrix(ws, "Корреляционная матрица", variables, result.get("matrix") or [])
            _append_matrix(ws, "P-values", variables, result.get("p_values") or [], format_p_value)
            _append_matrix(ws, "N matrix", variables, result.get("n_matrix") or [], format_value)
        elif section_type == "crosstab":
            _append_crosstab(ws, result.get("crosstab") or {})
        elif section_type == "chi_square":
            chi = result.get("chi_square") or {}
            cramers_v = result.get("cramers_v") or {}
            append_key_value(ws, "chi2", format_number(chi.get("chi2")))
            append_key_value(ws, "p_value", format_p_value(chi.get("p_value")))
            append_key_value(ws, "dof", chi.get("dof"))
            append_key_value(ws, "Cramer's V", format_number(cramers_v.get("cramers_v")))
            append_key_value(ws, "Интерпретация", cramers_v.get("interpretation"))
            append_key_value(ws, "N", cramers_v.get("n"))
            append_key_value(ws, "Число строк", cramers_v.get("rows"))
            append_key_value(ws, "Число столбцов", cramers_v.get("columns"))
            _append_crosstab(ws, result.get("crosstab") or {})
            _append_matrix(ws, "Expected values", [], chi.get("expected") or [])
        elif section_type == "correspondence_analysis":
            append_key_value(ws, "method", result.get("method"))
            append_key_value(ws, "n", result.get("n"))
            append_key_value(ws, "n_rows", result.get("n_rows"))
            append_key_value(ws, "n_columns", result.get("n_columns"))
            append_key_value(ws, "n_dimensions", result.get("n_dimensions"))
            append_key_value(ws, "total_inertia", format_number(result.get("total_inertia")))
            warnings = result.get("warnings") or []
            if warnings:
                _append_table(ws, ["Warnings"], [[warning] for warning in warnings])
            _append_crosstab(ws, result.get("crosstab") or {})
            _append_table(
                ws,
                ["Dimension", "Eigenvalue", "Explained inertia", "Percent"],
                [
                    [
                        item.get("dimension"),
                        format_number(item.get("eigenvalue")),
                        format_number(item.get("explained_inertia")),
                        f"{float(item.get('explained_inertia') or 0) * 100:.2f}%",
                    ]
                    for item in result.get("dimensions") or []
                ],
            )

            def append_ca_coordinates(title_text, points):
                dimensions = [item.get("dimension") for item in result.get("dimensions") or []]
                rows = []
                for point in points:
                    coordinates = {item.get("dimension"): item.get("value") for item in point.get("coordinates") or []}
                    contributions = {item.get("dimension"): item.get("value") for item in point.get("contributions") or []}
                    rows.append([
                        point.get("label") or point.get("value"),
                        format_number(point.get("mass")),
                        *[format_number(coordinates.get(dimension)) for dimension in dimensions],
                        *[format_number(contributions.get(dimension)) for dimension in dimensions],
                        format_number(point.get("cos2")),
                    ])
                _append_table(
                    ws,
                    ["Category", "Mass", *dimensions, *[f"Contribution {dimension}" for dimension in dimensions], "Cos2"],
                    rows,
                )

            append_section_title(ws, "Row coordinates")
            append_ca_coordinates("Row coordinates", result.get("row_coordinates") or [])
            append_section_title(ws, "Column coordinates")
            append_ca_coordinates("Column coordinates", result.get("column_coordinates") or [])
        elif section_type == "factor_analysis":
            append_key_value(ws, "method", result.get("method"))
            append_key_value(ws, "n", result.get("n"))
            append_key_value(ws, "n_variables", result.get("n_variables"))
            append_key_value(ws, "n_factors", result.get("n_factors"))
            append_key_value(ws, "rotation", result.get("rotation"))
            append_key_value(ws, "standardize", result.get("standardize"))
            append_key_value(ws, "cumulative_explained_variance", format_number(result.get("cumulative_explained_variance")))
            _append_table(
                ws,
                ["Component", "Eigenvalue"],
                [
                    [f"Component {index}", format_number(value)]
                    for index, value in enumerate(result.get("eigenvalues") or [], start=1)
                ],
            )
            _append_table(
                ws,
                ["Factor", "Value", "Percent"],
                [
                    [item.get("factor"), format_number(item.get("value")), f"{float(item.get('value') or 0) * 100:.2f}%"]
                    for item in result.get("explained_variance") or []
                ],
            )
            factors = [item.get("factor") for item in result.get("explained_variance") or []]
            loading_rows = []
            for item in result.get("loadings") or []:
                factors_by_name = {
                    factor.get("factor"): factor.get("loading")
                    for factor in item.get("factors") or []
                }
                loading_rows.append([
                    item.get("label") or item.get("variable"),
                    format_number(item.get("communality")),
                    *[format_number(factors_by_name.get(factor)) for factor in factors],
                ])
            _append_table(ws, ["Variable", "Communality", *factors], loading_rows)
            recommendations = result.get("factor_recommendations") or {}
            append_key_value(ws, "kaiser_n_factors", recommendations.get("kaiser_n_factors"))
            append_key_value(ws, "selected_n_factors", recommendations.get("selected_n_factors"))
            append_key_value(ws, "kaiser_message", recommendations.get("message"))
            kmo = result.get("kmo") or {}
            append_key_value(ws, "kmo_overall", format_number(kmo.get("overall")))
            append_key_value(ws, "kmo_interpretation", kmo.get("interpretation"))
            _append_table(
                ws,
                ["Variable", "KMO", "Interpretation"],
                [
                    [item.get("label") or item.get("code"), format_number(item.get("kmo")), item.get("interpretation")]
                    for item in kmo.get("variables") or []
                ],
            )
            bartlett = result.get("bartlett") or {}
            append_key_value(ws, "bartlett_chi_square", format_number(bartlett.get("chi_square")))
            append_key_value(ws, "bartlett_dof", bartlett.get("dof"))
            append_key_value(ws, "bartlett_p_value", format_p_value(bartlett.get("p_value")))
            append_key_value(ws, "bartlett_significant", bartlett.get("significant"))
            append_key_value(ws, "bartlett_interpretation", bartlett.get("interpretation"))
            _append_table(
                ws,
                ["Component", "Eigenvalue", "Explained variance", "Cumulative"],
                [
                    [
                        item.get("component"),
                        format_number(item.get("eigenvalue")),
                        format_number(item.get("explained_variance")),
                        format_number(item.get("cumulative_explained_variance")),
                    ]
                    for item in result.get("scree") or []
                ],
            )
            factor_scores = result.get("factor_scores") or []
            if factor_scores:
                score_factors = [score.get("factor") for score in (factor_scores[0].get("scores") or [])]
                _append_table(
                    ws,
                    ["response_id", *score_factors],
                    [
                        [
                            item.get("response_id"),
                            *[format_number(score.get("value")) for score in item.get("scores") or []],
                        ]
                        for item in factor_scores
                    ],
                )
            warnings = result.get("warnings") or []
            if warnings:
                _append_table(ws, ["Warnings"], [[warning] for warning in warnings])
            if kmo.get("warnings"):
                _append_table(ws, ["KMO warnings"], [[warning] for warning in kmo.get("warnings") or []])
            if bartlett.get("warnings"):
                _append_table(ws, ["Bartlett warnings"], [[warning] for warning in bartlett.get("warnings") or []])
        elif section_type == "cluster_analysis":
            append_key_value(ws, "method", result.get("method"))
            append_key_value(ws, "n", result.get("n"))
            append_key_value(ws, "n_clusters", result.get("n_clusters"))
            append_key_value(ws, "standardize", result.get("standardize"))
            append_key_value(ws, "inertia", format_number(result.get("inertia")))
            variables = result.get("variables") or []
            _append_table(
                ws,
                ["Cluster", "Size", "Percent", *[
                    variable.get("label") or variable.get("code")
                    for variable in variables
                ]],
                [
                    [
                        cluster.get("cluster"),
                        cluster.get("size"),
                        format_number(cluster.get("percent")),
                        *[
                            format_number((cluster.get("centroid") or {}).get(variable.get("code")))
                            for variable in variables
                        ],
                    ]
                    for cluster in result.get("clusters") or []
                ],
            )
            for profile in result.get("cluster_profiles") or []:
                cluster = profile.get("cluster")
                append_section_title(ws, f"Cluster {cluster} profile")
                append_key_value(ws, "size", profile.get("size"))
                append_key_value(ws, "percent", format_number(profile.get("percent")))
                append_key_value(ws, "interpretation", profile.get("interpretation"))
                _append_table(
                    ws,
                    ["Feature", "Type", "Cluster value", "Overall value", "Difference", "Score", "Interpretation"],
                    [
                        [
                            feature.get("label"),
                            feature.get("type"),
                            format_number(feature.get("cluster_value")),
                            format_number(feature.get("overall_value")),
                            format_number(feature.get("difference")),
                            format_number(feature.get("score")),
                            feature.get("interpretation"),
                        ]
                        for feature in profile.get("top_distinguishing_features") or []
                    ],
                )
                _append_table(
                    ws,
                    ["Variable", "Cluster mean", "Overall mean", "Difference", "z_difference", "Interpretation"],
                    [
                        [
                            item.get("label"),
                            format_number(item.get("cluster_mean")),
                            format_number(item.get("overall_mean")),
                            format_number(item.get("difference")),
                            format_number(item.get("z_difference")),
                            item.get("interpretation"),
                        ]
                        for item in profile.get("numeric_summary") or []
                    ],
                )
                _append_table(
                    ws,
                    ["Variable", "Cluster %", "Overall %", "Difference pp", "Interpretation"],
                    [
                        [
                            item.get("label"),
                            format_number(item.get("cluster_percent_selected")),
                            format_number(item.get("overall_percent_selected")),
                            format_number(item.get("difference_pp")),
                            item.get("interpretation"),
                        ]
                        for item in profile.get("binary_summary") or []
                    ],
                )
                categorical_rows = []
                for item in profile.get("categorical_summary") or []:
                    for category in item.get("categories") or []:
                        categorical_rows.append([
                            item.get("label"),
                            category.get("label"),
                            format_number(category.get("cluster_percent")),
                            format_number(category.get("overall_percent")),
                            format_number(category.get("difference_pp")),
                        ])
                _append_table(ws, ["Variable", "Category", "Cluster %", "Overall %", "Difference pp"], categorical_rows)
            warnings = result.get("warnings") or []
            if warnings:
                _append_table(ws, ["Warnings"], [[warning] for warning in warnings])
        elif section_type == "group_comparison":
            test = result.get("test") or {}
            effect_size = result.get("effect_size") or {}
            append_key_value(ws, "method", result.get("method_name") or result.get("method"))
            append_key_value(ws, "group variable", (result.get("group_variable") or {}).get("label"))
            append_key_value(ws, "value variable", (result.get("value_variable") or {}).get("label"))
            append_key_value(ws, "n", result.get("n"))
            append_key_value(ws, "n_groups", result.get("n_groups"))
            append_key_value(ws, "statistic", format_number(test.get("statistic")))
            append_key_value(ws, "p_value", format_p_value(test.get("p_value")))
            append_key_value(ws, "significant", test.get("significant"))
            append_key_value(ws, "interpretation", test.get("interpretation"))
            append_key_value(ws, "effect_size_type", effect_size.get("type"))
            append_key_value(ws, "effect_size_value", format_number(effect_size.get("value")))
            append_key_value(ws, "effect_size_interpretation", effect_size.get("interpretation"))
            _append_table(
                ws,
                ["Group", "n", "Mean", "Median", "Std", "Min", "Max"],
                [
                    [
                        group.get("label") or group.get("group"),
                        group.get("n"),
                        format_number(group.get("mean")),
                        format_number(group.get("median")),
                        format_number(group.get("std")),
                        format_number(group.get("min")),
                        format_number(group.get("max")),
                    ]
                    for group in result.get("groups") or []
                ],
            )
            warnings = result.get("warnings") or []
            if warnings:
                _append_table(ws, ["Warnings"], [[warning] for warning in warnings])
            post_hoc = result.get("post_hoc") or {}
            if post_hoc.get("enabled"):
                append_section_title(ws, "Post-hoc comparisons")
                append_key_value(ws, "post_hoc_method", post_hoc.get("method"))
                append_key_value(ws, "p_adjust", post_hoc.get("p_adjust"))
                append_key_value(ws, "comparisons_count", post_hoc.get("comparisons_count"))
                post_hoc_warnings = post_hoc.get("warnings") or []
                if post_hoc_warnings:
                    _append_table(ws, ["Post-hoc warnings"], [[warning] for warning in post_hoc_warnings])
                _append_table(
                    ws,
                    [
                        "Group A",
                        "Group B",
                        "Test",
                        "Statistic",
                        "p-value",
                        "Adjusted p-value",
                        "Significant",
                        "Difference",
                        "Effect size type",
                        "Effect size value",
                        "Effect size interpretation",
                    ],
                    [
                        [
                            comparison.get("group_a_label") or comparison.get("group_a"),
                            comparison.get("group_b_label") or comparison.get("group_b"),
                            comparison.get("test"),
                            format_number(comparison.get("statistic")),
                            format_p_value(comparison.get("p_value")),
                            format_p_value(comparison.get("p_adjusted")),
                            comparison.get("significant"),
                            format_number(comparison.get("difference")),
                            (comparison.get("effect_size") or {}).get("type"),
                            format_number((comparison.get("effect_size") or {}).get("value")),
                            (comparison.get("effect_size") or {}).get("interpretation"),
                        ]
                        for comparison in post_hoc.get("comparisons") or []
                    ],
                )
        elif section_type == "time_analysis":
            summary = result.get("summary") or {}
            for key in (
                "total_started",
                "total_finished",
                "total_completed",
                "total_screened_out",
                "total_active_unfinished",
                "completion_rate",
                "screenout_rate",
                "finish_rate",
                "average_completion_time_seconds",
                "median_completion_time_seconds",
                "min_completion_time_seconds",
                "max_completion_time_seconds",
                "average_screenout_time_seconds",
                "median_screenout_time_seconds",
                "min_screenout_time_seconds",
                "max_screenout_time_seconds",
            ):
                append_key_value(ws, key, format_number(summary.get(key)))
            _append_table(
                ws,
                ["Bucket", "Count", "Percent"],
                [
                    [item.get("label"), item.get("count"), format_number(item.get("percent"))]
                    for item in result.get("completion_time_distribution") or []
                ],
            )
            _append_table(
                ws,
                ["Screenout bucket", "Count", "Percent"],
                [
                    [item.get("label"), item.get("count"), format_number(item.get("percent"))]
                    for item in result.get("screenout_time_distribution") or []
                ],
            )
            _append_table(
                ws,
                ["Reason", "Count", "% screened out", "Avg time to screenout"],
                [
                    [
                        item.get("reason"),
                        item.get("count"),
                        format_number(item.get("percent_screened_out")),
                        format_number(item.get("average_time_to_screenout_seconds")),
                    ]
                    for item in result.get("screenout_reasons") or []
                ],
            )
            _append_table(
                ws,
                ["Group", "Started", "Completed", "Screened out", "Completion rate", "Screenout rate", "Median completion", "Median screenout"],
                [
                    [
                        item.get("group_label") or item.get("group"),
                        item.get("total_started"),
                        item.get("total_completed"),
                        item.get("total_screened_out"),
                        format_number(item.get("completion_rate")),
                        format_number(item.get("screenout_rate")),
                        format_number((item.get("completion_time") or {}).get("median")),
                        format_number((item.get("screenout_time") or {}).get("median")),
                    ]
                    for item in result.get("group_breakdown") or []
                ],
            )
            group_test = result.get("group_time_test") or {}
            if group_test:
                append_key_value(ws, "group_time_test_method", group_test.get("method"))
                append_key_value(ws, "group_time_test_statistic", format_number(group_test.get("statistic")))
                append_key_value(ws, "group_time_test_p_value", format_p_value(group_test.get("p_value")))
                append_key_value(ws, "group_time_test_significant", group_test.get("significant"))
                append_key_value(ws, "group_time_test_interpretation", group_test.get("interpretation"))
            warnings = result.get("warnings") or []
            if warnings:
                _append_table(ws, ["Warnings"], [[warning] for warning in warnings])
        elif section_type == "reliability_analysis":
            append_key_value(ws, "method", result.get("method"))
            append_key_value(ws, "n", result.get("n"))
            append_key_value(ws, "n_items", result.get("n_items"))
            append_key_value(ws, "alpha", format_number(result.get("alpha")))
            append_key_value(ws, "standardized_alpha", format_number(result.get("standardized_alpha")))
            append_key_value(ws, "mean_inter_item_correlation", format_number(result.get("mean_inter_item_correlation")))
            append_key_value(ws, "interpretation", result.get("interpretation"))
            warnings = result.get("warnings") or []
            if warnings:
                _append_table(ws, ["Warnings"], [[warning] for warning in warnings])
            _append_table(
                ws,
                ["Item", "Mean", "Variance", "Std", "Item-total correlation", "Alpha if deleted"],
                [
                    [
                        item.get("label") or item.get("code"),
                        format_number(item.get("mean")),
                        format_number(item.get("variance")),
                        format_number(item.get("std")),
                        format_number(item.get("item_total_correlation")),
                        format_number(item.get("alpha_if_deleted")),
                    ]
                    for item in result.get("item_statistics") or []
                ],
            )
            variables = result.get("variables") or []
            matrix = result.get("inter_item_correlation_matrix") or []
            if variables and matrix:
                labels = [variable.get("label") or variable.get("code") for variable in variables]
                _append_table(
                    ws,
                    ["Item", *labels],
                    [
                        [
                            labels[row_index],
                            *[
                                format_number(value)
                                for value in (matrix[row_index] if row_index < len(matrix) else [])
                            ],
                        ]
                        for row_index in range(len(labels))
                    ],
                )
        elif section_type == "scale_index":
            append_key_value(ws, "title", result.get("title"))
            append_key_value(ws, "calculation", result.get("calculation"))
            append_key_value(ws, "n_items", result.get("n_items"))
            append_key_value(ws, "n_scored", result.get("n_scored"))
            append_key_value(ws, "min_answered_items", result.get("min_answered_items"))
            append_key_value(ws, "n_complete_cases_for_alpha", result.get("n_complete_cases_for_alpha"))
            append_key_value(ws, "missing_count", result.get("missing_count"))
            summary = result.get("score_summary") or {}
            _append_table(
                ws,
                ["Metric", "Value"],
                [
                    ["n", summary.get("n")],
                    ["mean", format_number(summary.get("mean"))],
                    ["median", format_number(summary.get("median"))],
                    ["std", format_number(summary.get("std"))],
                    ["variance", format_number(summary.get("variance"))],
                    ["min", format_number(summary.get("min"))],
                    ["max", format_number(summary.get("max"))],
                    ["p25", format_number(summary.get("p25"))],
                    ["p75", format_number(summary.get("p75"))],
                    ["iqr", format_number(summary.get("iqr"))],
                ],
            )
            reliability = result.get("reliability") or {}
            if reliability:
                _append_table(
                    ws,
                    ["Reliability", "Value"],
                    [
                        ["alpha", format_number(reliability.get("alpha"))],
                        ["standardized_alpha", format_number(reliability.get("standardized_alpha"))],
                        ["mean_inter_item_correlation", format_number(reliability.get("mean_inter_item_correlation"))],
                        ["interpretation", reliability.get("interpretation")],
                    ],
                )
                if reliability.get("warnings"):
                    _append_table(ws, ["Reliability warnings"], [[warning] for warning in reliability.get("warnings") or []])
            _append_table(
                ws,
                ["Item", "Reverse", "Min value", "Max value", "n", "Missing", "Mean", "Std", "Min", "Max", "Item-total correlation"],
                [
                    [
                        item.get("label") or item.get("code"),
                        item.get("reverse"),
                        item.get("min_value"),
                        item.get("max_value"),
                        item.get("n"),
                        item.get("missing"),
                        format_number(item.get("mean")),
                        format_number(item.get("std")),
                        format_number(item.get("min")),
                        format_number(item.get("max")),
                        format_number(item.get("item_total_correlation")),
                    ]
                    for item in result.get("item_statistics") or []
                ],
            )
            _append_table(
                ws,
                ["Bucket", "Count", "Percent"],
                [
                    [item.get("label"), item.get("count"), format_number(item.get("percent"))]
                    for item in result.get("score_distribution") or []
                ],
            )
            _append_table(
                ws,
                ["Response ID", "Score", "Answered items", "Missing items"],
                [
                    [
                        item.get("response_id"),
                        format_number(item.get("score")),
                        item.get("answered_items"),
                        item.get("missing_items"),
                    ]
                    for item in result.get("scores") or []
                ],
            )
            if result.get("warnings"):
                _append_table(ws, ["Warnings"], [[warning] for warning in result.get("warnings") or []])
        elif section_type == "missing_analysis":
            summary = result.get("summary") or {}
            append_key_value(ws, "total_completed_normal", summary.get("total_completed_normal"))
            append_key_value(ws, "questions_count", summary.get("questions_count"))
            append_key_value(ws, "total_shown_slots", summary.get("total_shown_slots"))
            append_key_value(ws, "total_answered_slots", summary.get("total_answered_slots"))
            append_key_value(ws, "total_skipped_slots", summary.get("total_skipped_slots"))
            append_key_value(ws, "total_not_shown_slots", summary.get("total_not_shown_slots"))
            append_key_value(ws, "overall_skip_rate_shown", format_number(summary.get("overall_skip_rate_shown")))
            append_key_value(ws, "overall_visibility_rate", format_number(summary.get("overall_visibility_rate")))
            append_key_value(ws, "questions_with_high_missing", summary.get("questions_with_high_missing"))
            append_key_value(ws, "questions_with_moderate_missing", summary.get("questions_with_moderate_missing"))
            append_key_value(ws, "questions_with_low_visibility", summary.get("questions_with_low_visibility"))
            append_key_value(ws, "note", "Вопросы, не показанные из-за ветвления, не учитываются как реальные пропуски.")

            _append_table(
                ws,
                [
                    "Question",
                    "Type",
                    "Required",
                    "Page",
                    "Shown",
                    "Not shown",
                    "Answered",
                    "Skipped",
                    "Visibility rate",
                    "Answer rate shown",
                    "Skip rate shown",
                    "Answer rate total",
                    "Missing type",
                    "Interpretation",
                ],
                [
                    [
                        item.get("label"),
                        item.get("qtype"),
                        item.get("required"),
                        item.get("page_title"),
                        item.get("shown_count"),
                        item.get("not_shown_count"),
                        item.get("answered_count"),
                        item.get("skipped_count"),
                        format_number(item.get("visibility_rate")),
                        format_number(item.get("answer_rate_shown")),
                        format_number(item.get("skip_rate_shown")),
                        format_number(item.get("answer_rate_total")),
                        item.get("missing_type"),
                        item.get("interpretation"),
                    ]
                    for item in result.get("questions") or []
                ],
            )
            _append_table(
                ws,
                ["Top skipped", "Shown", "Skipped", "Skip rate shown", "Visibility rate"],
                [
                    [
                        item.get("label"),
                        item.get("shown_count"),
                        item.get("skipped_count"),
                        format_number(item.get("skip_rate_shown")),
                        format_number(item.get("visibility_rate")),
                    ]
                    for item in result.get("top_skipped_questions") or []
                ],
            )
            _append_table(
                ws,
                ["Low visibility", "Shown", "Skipped", "Skip rate shown", "Visibility rate"],
                [
                    [
                        item.get("label"),
                        item.get("shown_count"),
                        item.get("skipped_count"),
                        format_number(item.get("skip_rate_shown")),
                        format_number(item.get("visibility_rate")),
                    ]
                    for item in result.get("low_visibility_questions") or []
                ],
            )
            _append_table(
                ws,
                ["Required with missing", "Shown", "Skipped", "Skip rate shown", "Visibility rate"],
                [
                    [
                        item.get("label"),
                        item.get("shown_count"),
                        item.get("skipped_count"),
                        format_number(item.get("skip_rate_shown")),
                        format_number(item.get("visibility_rate")),
                    ]
                    for item in result.get("required_questions_with_missing") or []
                ],
            )
            if result.get("groups"):
                _append_table(
                    ws,
                    ["Group", "Shown slots", "Answered slots", "Skipped slots", "Skip rate shown"],
                    [
                        [
                            group.get("group_label"),
                            group.get("total_shown_slots"),
                            group.get("total_answered_slots"),
                            group.get("total_skipped_slots"),
                            format_number(group.get("overall_skip_rate_shown")),
                        ]
                        for group in result.get("groups") or []
                    ],
                )
            screened_out = result.get("screened_out_context") or {}
            if screened_out:
                _append_table(
                    ws,
                    ["Screened out context", "Value"],
                    [
                        ["total_screened_out", screened_out.get("total_screened_out")],
                        ["average_seen_questions_before_screenout", format_number(screened_out.get("average_seen_questions_before_screenout"))],
                        ["note", screened_out.get("note")],
                    ],
                )
            if result.get("warnings"):
                _append_table(ws, ["Warnings"], [[warning] for warning in result.get("warnings") or []])
        elif section_type == "regression":
            append_key_value(ws, "target", get_variable_label(result, result.get("target")))
            append_key_value(ws, "features", ", ".join(get_variable_label(result, code) for code in (result.get("features") or [])))
            append_key_value(ws, "n", result.get("n"))
            append_key_value(ws, "r2", format_number(result.get("r2")))
            append_key_value(ws, "adjusted_r2", format_number(result.get("adjusted_r2")))
            _append_table(
                ws,
                ["Коэффициент", "Значение"],
                [
                    [get_variable_label(result, coefficient.get("name")), format_number(coefficient.get("value"))]
                    for coefficient in result.get("coefficients") or []
                ],
            )
        elif section_type == "logistic_regression":
            metrics = result.get("metrics") or {}
            confusion = result.get("confusion_matrix") or {}
            append_key_value(ws, "method", result.get("method"))
            append_key_value(ws, "target", get_variable_label(result, result.get("target")))
            append_key_value(ws, "features", ", ".join(get_variable_label(result, code) for code in (result.get("features") or [])))
            append_key_value(ws, "n", result.get("n"))
            append_key_value(ws, "positive_class_count", result.get("positive_class_count"))
            append_key_value(ws, "negative_class_count", result.get("negative_class_count"))
            append_key_value(ws, "base_rate", format_number(result.get("base_rate")))
            append_key_value(ws, "threshold", format_number(result.get("threshold")))
            append_key_value(ws, "accuracy", format_number(metrics.get("accuracy")))
            append_key_value(ws, "precision", format_number(metrics.get("precision")))
            append_key_value(ws, "recall", format_number(metrics.get("recall")))
            append_key_value(ws, "f1", format_number(metrics.get("f1")))
            append_key_value(ws, "mcfadden_r2", format_number(metrics.get("mcfadden_r2")))
            _append_table(
                ws,
                ["", "Predicted 0", "Predicted 1"],
                [
                    ["Actual 0", confusion.get("tn"), confusion.get("fp")],
                    ["Actual 1", confusion.get("fn"), confusion.get("tp")],
                ],
            )
            _append_table(
                ws,
                ["Variable", "Coefficient", "Odds ratio", "Interpretation"],
                [
                    [
                        get_variable_label(result, coefficient.get("name")),
                        format_number(coefficient.get("coefficient")),
                        format_number(coefficient.get("odds_ratio")),
                        coefficient.get("interpretation"),
                    ]
                    for coefficient in result.get("coefficients") or []
                ],
            )
            for warning in result.get("warnings") or []:
                append_key_value(ws, "warning", warning)
        ws.append([])


def _add_missing_analysis_sheet(ws, section):
    result = section.get("result") or {}
    title = section.get("title") or "Анализ пропусков"
    summary = result.get("summary") or {}
    append_section_title(ws, title)
    append_key_value(ws, "total_completed_normal", summary.get("total_completed_normal"))
    append_key_value(ws, "questions_count", summary.get("questions_count"))
    append_key_value(ws, "total_shown_slots", summary.get("total_shown_slots"))
    append_key_value(ws, "total_answered_slots", summary.get("total_answered_slots"))
    append_key_value(ws, "total_skipped_slots", summary.get("total_skipped_slots"))
    append_key_value(ws, "total_not_shown_slots", summary.get("total_not_shown_slots"))
    append_key_value(ws, "overall_skip_rate_shown", format_number(summary.get("overall_skip_rate_shown")))
    append_key_value(ws, "overall_visibility_rate", format_number(summary.get("overall_visibility_rate")))
    append_key_value(ws, "note", "Вопросы, не показанные из-за ветвления, не учитываются как реальные пропуски.")
    ws.append([])
    _append_table(
        ws,
        ["Question", "Type", "Required", "Shown", "Not shown", "Answered", "Skipped", "Skip rate shown", "Visibility rate", "Missing type", "Interpretation"],
        [
            [
                item.get("label"),
                item.get("qtype"),
                item.get("required"),
                item.get("shown_count"),
                item.get("not_shown_count"),
                item.get("answered_count"),
                item.get("skipped_count"),
                format_number(item.get("skip_rate_shown")),
                format_number(item.get("visibility_rate")),
                item.get("missing_type"),
                item.get("interpretation"),
            ]
            for item in result.get("questions") or []
        ],
    )
    _append_table(
        ws,
        ["Top skipped", "Shown", "Skipped", "Skip rate shown", "Visibility rate"],
        [[item.get("label"), item.get("shown_count"), item.get("skipped_count"), format_number(item.get("skip_rate_shown")), format_number(item.get("visibility_rate"))] for item in result.get("top_skipped_questions") or []],
    )
    _append_table(
        ws,
        ["Low visibility", "Shown", "Skipped", "Skip rate shown", "Visibility rate"],
        [[item.get("label"), item.get("shown_count"), item.get("skipped_count"), format_number(item.get("skip_rate_shown")), format_number(item.get("visibility_rate"))] for item in result.get("low_visibility_questions") or []],
    )
    _append_table(
        ws,
        ["Required with missing", "Shown", "Skipped", "Skip rate shown", "Visibility rate"],
        [[item.get("label"), item.get("shown_count"), item.get("skipped_count"), format_number(item.get("skip_rate_shown")), format_number(item.get("visibility_rate"))] for item in result.get("required_questions_with_missing") or []],
    )
    if result.get("groups"):
        _append_table(
            ws,
            ["Group", "Shown slots", "Answered slots", "Skipped slots", "Skip rate shown"],
            [[group.get("group_label"), group.get("total_shown_slots"), group.get("total_answered_slots"), group.get("total_skipped_slots"), format_number(group.get("overall_skip_rate_shown"))] for group in result.get("groups") or []],
        )
    screened_out = result.get("screened_out_context") or {}
    if screened_out:
        _append_table(
            ws,
            ["Screened out context", "Value"],
            [
                ["total_screened_out", screened_out.get("total_screened_out")],
                ["average_seen_questions_before_screenout", format_number(screened_out.get("average_seen_questions_before_screenout"))],
                ["note", screened_out.get("note")],
            ],
        )


def build_analytics_xlsx(survey, analytic_result, analysis_report) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:  # pragma: no cover - depends on deployment environment
        raise RuntimeError("XLSX export requires openpyxl to be installed.") from exc

    wb = Workbook()
    ws_survey = wb.active
    ws_survey.title = "Опрос"
    ws_summary = wb.create_sheet("Сводка")
    ws_screening = wb.create_sheet("Скрининг")
    ws_questions = wb.create_sheet("Вопросы")
    ws_report = wb.create_sheet("Отчёт")

    analytic_data = analytic_result.data or {}

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    ws_survey.append(["Показатель", "Значение"])
    _style_header(ws_survey, 1)
    for key, value in (
        ("ID опроса", survey.id),
        ("Название", survey.title),
        ("Описание", survey.description),
        ("Статус", survey.status),
        ("Дата начала", format_datetime(survey.starts_at)),
        ("Дата окончания", format_datetime(survey.ends_at)),
        ("Анонимный", "Да" if survey.is_anonymous else "Нет"),
        ("Дата формирования XLSX", format_datetime(timezone.now())),
        ("Название среза общей аналитики", analytic_result.title),
        ("Дата среза", format_datetime(analytic_result.generated_at)),
        ("Название пользовательского отчёта", analysis_report.title),
        ("Дата пользовательского отчёта", format_datetime(analysis_report.created_at)),
    ):
        append_key_value(ws_survey, key, value)

    ws_summary.append(["Показатель", "Значение"])
    _style_header(ws_summary, 1)
    summary = analytic_data.get("summary") or {}
    summary_labels = {
        "total_started": "Начали",
        "total_completed": "Завершили полностью",
        "total_screened_out": "Отсечены",
        "total_finished": "Завершили всего",
        "completion_rate": "Процент полного завершения",
        "screenout_rate": "Процент скрининга",
        "finish_rate": "Процент завершения всего",
        "average_completion_time": "Среднее время завершения",
        "average_screenout_time": "Среднее время до скрининга",
        "questions_count": "Количество вопросов",
    }
    for key, label in summary_labels.items():
        append_key_value(ws_summary, label, format_number(summary.get(key)))

    ws_screening.append(["Показатель", "Значение"])
    _style_header(ws_screening, 1)
    screening = analytic_data.get("screening") or {}
    append_key_value(ws_screening, "total_screened_out", screening.get("total_screened_out"))
    append_key_value(ws_screening, "average_screenout_time", format_number(screening.get("average_screenout_time")))
    ws_screening.append([])
    reasons = screening.get("reasons") or []
    if reasons:
        _append_table(ws_screening, ["Причина", "Количество"], [[item.get("reason"), item.get("count")] for item in reasons])
    else:
        ws_screening.append(["Нет screened_out ответов"])

    _add_questions_sheet(ws_questions, survey, analytic_data)
    _add_report_sheet(ws_report, analysis_report)

    report_result = _report_result(analysis_report)
    missing_sections = [
        section
        for section in report_result.get("sections") or []
        if section.get("type") == "missing_analysis" and not section.get("error")
    ]
    for index, section in enumerate(missing_sections, start=1):
        sheet_title = "Анализ пропусков" if index == 1 else f"Анализ пропусков {index}"
        ws_missing = wb.create_sheet(sheet_title[:31])
        ws_missing.freeze_panes = "A2"
        _add_missing_analysis_sheet(ws_missing, section)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True
        for cell in ws[1]:
            cell.font = Font(bold=True)
        autosize_columns(ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
